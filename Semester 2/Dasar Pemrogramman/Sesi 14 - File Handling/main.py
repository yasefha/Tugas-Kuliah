from openpyxl import load_workbook

# buka file excel
wb = load_workbook('data.xlsx')
ws = wb.active

# hitung jumlah baris terakhir
max_row = ws.max_row

# print data
print('Data Sebelum Penambahan:')
for row in ws.iter_rows(min_row=max_row-1, max_row=max_row, values_only=True):
    print(row)

# tambah data baru
nama_baru = 'Naura Avantika Rinjani'
alamat_baru = 'Bandung'
ws.append([nama_baru, alamat_baru])

# save perubahan
wb.save('data.xlsx')

max_row = ws.max_row

print()
print('Data Setelah Penambahan:')
for row in ws.iter_rows(min_row=max_row-1, max_row=max_row, values_only=True):
    print(row)